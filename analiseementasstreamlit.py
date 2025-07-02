# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import re
import os
import zipfile
import tempfile
import pdfplumber
from io import BytesIO
from sentence_transformers import SentenceTransformer, util
from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import CountVectorizer
from nltk.corpus import stopwords
import openai


# Configuração da página
st.set_page_config(layout="wide")
st.title("📂📑 Análise de Ementas via pasta .zip")

# --- 1) Upload do ZIP de PDFs de ementas ---
uploaded_zip = st.file_uploader(
    "Faça upload dos arquivos, em PDF das ementas, em uma pasta .zip",
    type=["zip"]
)
if not uploaded_zip:
    st.info("Aguardando upload do ZIP...")
    st.stop()

# --- 2) Extrai e processa os PDFs ---
with tempfile.TemporaryDirectory() as tmpdir:
    z = zipfile.ZipFile(uploaded_zip)
    z.extractall(tmpdir)

    registros = []
    for root, _, files in os.walk(tmpdir):
        for fn in files:
            if not fn.lower().endswith(".pdf"):
                continue
            path = os.path.join(root, fn)
            texto = ""
            with pdfplumber.open(path) as pdf:
                for p in pdf.pages:
                    texto += (p.extract_text() or "") + "\n"

            # --- Limpeza: remove linhas de paginação tipo "2 de 3", "10 de 12" etc. ---
            texto = re.sub(r"(?m)^\s*\d+\s+de\s+\d+\s*$", "", texto)

            # extrai nome e código
            m = re.search(
                r"UNIDADE CURRICULAR[:\s]*(.+?)\s*\(\s*(\d+)\s*\)",
                texto, re.IGNORECASE | re.DOTALL
            )
            nome = m.group(1).strip() if m else fn
            cod  = m.group(2).strip() if m else fn

            # extrai conteúdo programático
            m2 = re.search(
                r"Conte[úu]do program[aá]tico\s*[:\-–]?\s*(.*?)(?=\n\s*Bibliografia|\Z)",
                texto, re.IGNORECASE | re.DOTALL
            )
            conteudo = m2.group(1).strip() if m2 else ""

            registros.append({
                "COD_EMENTA": cod,
                "NOME UC": nome,
                "CONTEUDO_PROGRAMATICO": conteudo
            })

    df_ementas = pd.DataFrame(registros)

# --- Pergunta ao usuário sobre correção de pontuação ---
st.success(f"{len(df_ementas)} ementas carregadas.")

usar_gpt = st.checkbox(
    "Corrigir pontuação das ementas via OpenAI GPT antes da separação de frases? Utilizaremos o GPT3.5-Turbo."
)

if usar_gpt:
    api_key = st.text_input("Insira sua OpenAI API Key:", type="password")
    if api_key:
        # Usa a nova interface do SDK v1
        from openai import OpenAI
        client = OpenAI(api_key=api_key)

        @st.cache_data
        def corrigir_pontuacao(texto: str) -> str:
            prompt = (
                "Revise o texto abaixo para melhorar a pontuação: "
                "adicione ou ajuste pontos finais, vírgulas e demais sinais, "
                "mantendo o sentido original.\n\n"
                f"Texto:\n{texto}"
            )
            resp = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "Você é um especialista em revisão de texto."},
                    {"role": "user",   "content": prompt}
                ],
                temperature=0.0,
                max_tokens=len(texto.split()) // 2
            )
            return resp.choices[0].message.content.strip()

        with st.spinner("Corrigindo pontuação via GPT…"):
            df_ementas['CONTEUDO_PROGRAMATICO'] = (
                df_ementas['CONTEUDO_PROGRAMATICO']
                .apply(corrigir_pontuacao)
            )
        st.success("Pontuação corrigida em todas as ementas.")
    else:
        st.info("Para usar a correção via GPT, insira sua OpenAI API Key acima.")
else:
    st.info("Seguindo com a separação padrão de frases sem correção de pontuação.")

# --- Continuação do fluxo normal ---
st.subheader("Preview das primeiras ementas")
st.dataframe(df_ementas.head(5))

# --- 3) Upload do Excel ENADE ---
uploaded_enade = st.file_uploader(
    "Faça upload do Excel de competências ENADE",
    type=["xlsx"], key="enade"
)
if not uploaded_enade:
    st.info("Envie o arquivo ENADE para prosseguir.")
    st.stop()

enade = pd.read_excel(uploaded_enade).dropna(subset=['DESCRIÇÃO'])
# explode frases ENADE
enade['FRASE_ENADE'] = (
    enade['DESCRIÇÃO']
    .str.replace('\n',' ')
    .str.split(r'[.;]')
)
enade_expl = (
    enade.explode('FRASE_ENADE')
         .assign(FRASE_ENADE=lambda df: df['FRASE_ENADE'].str.strip())
)
enade_expl = enade_expl[enade_expl['FRASE_ENADE'].str.len() > 5].reset_index(drop=True)

# --- 4) Escolha da análise ---
analise = st.sidebar.selectbox("Escolha a Análise", [
    "Clusterização Ementas",
    "Matriz de Similaridade",
    "Matriz de Redundância",
    "Análise Ementa vs ENADE"
])

# --- 5) Carrega modelo em cache ---
@st.cache_resource
def load_model():
    return SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
model = load_model()


# --- 6A) Clusterização Ementas via KMeans + t-SNE c/ nomes automáticos GPT v1.0+ ---
if analise == "Clusterização Ementas":
    st.header("Clusterização das UCs")

    # 1) Agrega cada UC em um único texto
    df_group = (
        df_ementas
        .groupby(['COD_EMENTA','NOME UC'])['CONTEUDO_PROGRAMATICO']
        .apply(lambda texts: " ".join(texts))
        .reset_index()
    )

    # 2) Gera embeddings SBERT
    texts = df_group['CONTEUDO_PROGRAMATICO'].tolist()
    emb   = model.encode(texts, convert_to_tensor=False)

    # 3) Slider para número de clusters K
    max_k = min(10, len(emb))
    k = st.slider(
        "Número de clusters (K)",
        min_value=2,
        max_value=max_k,
        value=min(4, max_k),
        step=1
    )

    # 4) Executa KMeans
    kmeans = KMeans(n_clusters=k, random_state=42)
    df_group['cluster'] = kmeans.fit_predict(emb)

    # 5) Pergunta se quer usar GPT para nomear
    use_ai = st.radio(
        "Usar ChatGPT para nomear clusters?",
        ("Não, usar centróide", "Sim, usar GPT-3.5")
    )

    cluster_names = {}
    if use_ai == "Sim, usar GPT-3.5":
        key = st.text_input("Insira sua OpenAI API Key:", type="password")
        if key:
            openai.api_key = key
            for cid in range(k):
                exemplos = df_group[df_group['cluster']==cid]['CONTEUDO_PROGRAMATICO'].tolist()[:5]
                prompt = (
                    "Estas são ementas de um mesmo grupo de disciplinas:\n\n"
                    + "\n".join(f"- {e}" for e in exemplos)
                    + "\n\nPor favor, dê um nome curto (até 3 palavras) que resuma o tema comum."
                )
                try:
                    resp = openai.chat.completions.create(
                        model="gpt-3.5-turbo",
                        messages=[
                            {"role":"system","content":"Você resume grupos de ementas em um nome curto."},
                            {"role":"user",  "content":prompt}
                        ],
                        temperature=0.0,
                        max_tokens=20
                    )
                    nome = resp.choices[0].message.content.strip().strip('"')
                except Exception as e:
                    st.warning(f"GPT erro no cluster {cid}: {e}")
                    nome = f"Cluster {cid}"
                cluster_names[cid] = nome
        else:
            st.info("Informe a API Key para gerar nomes via GPT.")
            for cid in range(k):
                cluster_names[cid] = f"Cluster {cid}"
    else:
        # fallback centróide
        centroids = kmeans.cluster_centers_
        for cid in range(k):
            mask    = df_group['cluster']==cid
            idxs    = df_group[mask].index.to_numpy()
            dists   = np.linalg.norm(emb[idxs] - centroids[cid], axis=1)
            rep_idx = idxs[dists.argmin()]
            cluster_names[cid] = df_group.at[rep_idx, 'NOME UC']

    df_group['cluster_name'] = df_group['cluster'].map(cluster_names)

    # 6) t-SNE
    coords = TSNE(n_components=2, random_state=42,
                  perplexity=min(30, max(2, len(emb)//3))
    ).fit_transform(emb)
    df_group['X'], df_group['Y'] = coords[:,0], coords[:,1]

    # 7) Plot
    fig, ax = plt.subplots(figsize=(8,6))
    palette = plt.cm.get_cmap("tab10", k)
    for cid in range(k):
        sub = df_group[df_group['cluster']==cid]
        ax.scatter(sub['X'], sub['Y'],
                   color=palette(cid),
                   label=cluster_names[cid],
                   s=40, alpha=0.7)
        for _, row in sub.iterrows():
            ax.text(row['X']+0.3, row['Y']+0.3,
                    str(row['COD_EMENTA']), fontsize=6)
    ax.set_xlabel("Dimensão 1 (t-SNE 1)")
    ax.set_ylabel("Dimensão 2 (t-SNE 2)")
    ax.set_title(f"Clusterização das Ementas - k = {k}", fontsize=14)
    ax.legend(title="Nome do Cluster", bbox_to_anchor=(1,1))
    st.pyplot(fig)

    # 8) Download gráfico
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=300, bbox_inches="tight")
    buf.seek(0)
    st.download_button("📥 Baixar Gráfico t-SNE",
                       buf, "tsne_kmeans.png", "image/png")

    # 9) Tabela de clusters
    st.subheader("Clusters atribuídos por UC")
    st.dataframe(df_group[['COD_EMENTA','NOME UC','cluster','cluster_name']])
    buf2 = BytesIO()
    df_group[['COD_EMENTA','NOME UC','cluster','cluster_name']] \
        .to_excel(buf2, index=False)
    buf2.seek(0)
    st.download_button("📥 Baixar Tabela de Clusters",
                       buf2, "clusters_ucs.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- 6B) Matriz de Similaridade ---
elif analise == "Matriz de Similaridade":
    st.header("Matriz de Similaridade")
    # explode ementas em frases
    ementa_expl = (
        df_ementas
        .assign(
            FRASE=lambda df: df['CONTEUDO_PROGRAMATICO']
                .str.replace('\n',' ')
                .str.split(r'[.;]')
        )
        .explode('FRASE')
        .assign(FRASE=lambda df: df['FRASE'].str.strip())
    )
    ementa_expl = ementa_expl[ementa_expl['FRASE'].str.len()>5]

    with st.spinner("Gerando embeddings…"):
        emb_e = model.encode(ementa_expl['FRASE'].tolist(), convert_to_tensor=True)
        emb_n = model.encode(enade_expl['FRASE_ENADE'].tolist(), convert_to_tensor=True)

    sim = util.cos_sim(emb_n, emb_e).cpu().numpy()
    rec = []
    idxs = ementa_expl.groupby('COD_EMENTA').indices
    for cod, sidx in idxs.items():
        for i,row in enade_expl.iterrows():
            rec.append({
                "COD_EMENTA": cod,
                "FRASE_ENADE": row['FRASE_ENADE'],
                "MAX_SIM": float(sim[i, sidx].max())
            })
    df_sim = (
        pd.DataFrame(rec)
        .pivot(index='COD_EMENTA', columns='FRASE_ENADE', values='MAX_SIM')
        .fillna(0)
    )

    # Exibe no Streamlit
    st.dataframe(df_sim.style.background_gradient(cmap="RdYlGn"))

    # 1) Grava DataFrame num buffer
    buf = BytesIO()
    df_sim.to_excel(buf, index=True, sheet_name="Similaridade")
    buf.seek(0)

    # 2) Carrega workbook do buffer
    wb = load_workbook(buf)
    ws = wb["Similaridade"]

    # 3) Define intervalo de células (coluna B até a última)
    min_col = 2
    max_col = ws.max_column
    max_row = ws.max_row
    col_letter = lambda idx: ws.cell(row=1, column=idx).column_letter
    range_str = f"{col_letter(min_col)}2:{col_letter(max_col)}{max_row}"

    # 4) Cria regra: vermelho em 0 → amarelo em 50% → verde em 1
    rule = ColorScaleRule(
        start_type='min', start_color='FF0000',
        mid_type='percentile', mid_value=50, mid_color='FFFF00',
        end_type='max', end_color='00FF00'
    )
    ws.conditional_formatting.add(range_str, rule)

    # 5) Salva de volta no buffer
    buf2 = BytesIO()
    wb.save(buf2)
    buf2.seek(0)

    # 6) Botão de download
    st.download_button(
        "⬇️ Baixar Matriz de Similaridade",
        data=buf2,
        file_name="sim_enade_ementa_colorido.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# --- 6C) Matriz de Redundância ---
elif analise == "Matriz de Redundância":
    st.header("Matriz de Redundância")
    df_group = (
        df_ementas
        .groupby('COD_EMENTA')['CONTEUDO_PROGRAMATICO']
        .apply(lambda txts: " ".join(txts))
        .reset_index()
    )
    emb = model.encode(df_group['CONTEUDO_PROGRAMATICO'].tolist(), convert_to_tensor=True)
    sim = util.cos_sim(emb, emb).cpu().numpy()
    df_red = pd.DataFrame(sim, index=df_group['COD_EMENTA'], columns=df_group['COD_EMENTA'])

    # Exibe no Streamlit
    st.dataframe(df_red.style.background_gradient(cmap="RdYlGn_r"))

    # 1) Grava DataFrame num buffer
    buf = BytesIO()
    df_red.to_excel(buf, index=True, sheet_name="Redundância")
    buf.seek(0)

    # 2) Carrega workbook do buffer
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(buf)
    ws = wb["Redundância"]

    # 3) Define intervalo de células (da coluna B até a última)
    min_col = 2
    max_col = ws.max_column
    max_row = ws.max_row
    # Função auxiliar para converter índice em letra de coluna
    def col_letter(idx):
        return ws.cell(row=1, column=idx).column_letter

    range_str = f"{col_letter(min_col)}2:{col_letter(max_col)}{max_row}"

    # 4) Cria regra: vermelho em valor mínimo → amarelo em 50% → verde em valor máximo
    rule = ColorScaleRule(
        start_type='min',      start_color='00FF00',  # verde
        mid_type='percentile', mid_value=50,          mid_color='FFFF00',  # amarelo
        end_type='max',        end_color='FF0000'      # vermelho
    )
    ws.conditional_formatting.add(range_str, rule)

    # 5) Salva de volta num novo buffer
    buf2 = BytesIO()
    wb.save(buf2)
    buf2.seek(0)

    # 6) Botão de download com o arquivo já colorido
    st.download_button(
        "⬇️ Baixar Matriz de Redundância",
        data=buf2,
        file_name="redundancia_uc_colorida.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- 6D) Análise Ementa vs ENADE (ajustado para usar a FRASE correta) ---
else:
    st.header("Análise Ementa vs ENADE")

    # 6D.1) Explode contextualizado em frases
    df_ctx = df_ementas.rename(columns={"CONTEUDO_PROGRAMATICO":"CONTEUDO_PROGRAMATICO"})
    df_ctx['FRASE'] = (
        df_ctx['CONTEUDO_PROGRAMATICO']
          .str.replace('\n',' ')
          .str.split(r'\.')
    )
    df_ctx = (
        df_ctx
        .explode('FRASE')
        .assign(FRASE=lambda d: d['FRASE'].str.strip())
    )
    df_ctx = df_ctx[df_ctx['FRASE'].str.len()>5].reset_index(drop=True)

    # 6D.2) Embeddings
    limiar = st.slider("Limiar de similaridade", 0.0, 1.0, 0.6, step=0.05)
    with st.spinner("Calculando embeddings..."):
        emb_f = model.encode(df_ctx['FRASE'].tolist(), convert_to_tensor=True)
        emb_n = model.encode(enade_expl['FRASE_ENADE'].tolist(), convert_to_tensor=True)
    simm = util.cos_sim(emb_n, emb_f).cpu().numpy()

    # 6D.3) Construção dos resultados
    records = []
    for i, row_enade in enade_expl.iterrows():
        sims    = simm[i]                # similaridades contra todas as FRASEs
        max_sim = float(sims.max())
        idx_max = int(sims.argmax())
        cod_max = df_ctx.loc[idx_max, 'COD_EMENTA']
        # **Aqui pegamos a FRASE, não o bloco inteiro!**
        texto_max = df_ctx.loc[idx_max, 'FRASE']
        acima     = df_ctx.loc[sims >= limiar, 'COD_EMENTA'].unique().tolist()

        records.append({
            "FRASE_ENADE":     row_enade['FRASE_ENADE'],
            "DIMENSÃO":        row_enade['DIMENSAO'],
            "MAX_SIM":         round(max_sim, 3),
            "COD_EMENTA_MAX":  cod_max,
            "TEXTO_MAX":       texto_max,
            f"UCs_>={int(limiar*100)}%": "; ".join(map(str, acima))
        })

    # 6D.4) Exibe e oferece download
    df_res = pd.DataFrame(records)
    st.dataframe(df_res)
    buf = BytesIO()
    df_res.to_excel(buf, index=False)
    buf.seek(0)
    st.download_button(
        "📥 Baixar Análise Expandida vs ENADE",
        data=buf,
        file_name="analise_ementa_expandida_vs_enade.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

